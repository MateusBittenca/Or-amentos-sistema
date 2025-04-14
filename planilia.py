import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
from PIL import Image
import pytesseract
import re
import sys

class ComprovantesManager:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        # Carregar o arquivo Excel, mantendo a formatação existente
        self.workbook = openpyxl.load_workbook(excel_path)
        self.sheet = self.workbook.active
        
        # Definir cores para formatação
        self.verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Carregar dados para pandas para facilitar a manipulação
        self.df = pd.read_excel(excel_path)
    
    def preencher_pagamento(self, valor_str, atividade, pagador, setor=None):
        """
        Preenche a tabela com os dados do comprovante
        
        Args:
            valor_str (str): Valor do pagamento extraído do comprovante (formato: R$ X.XXX,XX)
            atividade (str): Descrição da atividade
            pagador (str): Quem pagou (Alex-Rute ou Diego-Ana)
            setor (str, optional): Setor relacionado ao pagamento
        """
        # Converter o valor de string para float
        valor_str = valor_str.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
        try:
            valor = float(valor_str)
        except ValueError:
            print(f"Erro ao converter valor: {valor_str}")
            return False
        
        # Procurar linha com a atividade especificada
        linhas_encontradas = self.df[self.df['ATIVIDADE'].str.strip().str.lower() == atividade.strip().lower()].index.tolist()
        
        # Se setor for especificado, filtrar também pelo setor
        if setor and len(linhas_encontradas) > 1:
            linhas_com_setor = self.df[(self.df['ATIVIDADE'].str.strip().str.lower() == atividade.strip().lower()) & 
                                      (self.df['SETOR'].str.strip().str.lower() == setor.strip().lower())].index.tolist()
            if linhas_com_setor:
                linhas_encontradas = linhas_com_setor
        
        if not linhas_encontradas:
            print(f"Atividade '{atividade}' não encontrada na tabela.")
            return False
        
        # Para simplificar, vamos usar a primeira linha encontrada
        linha_excel = linhas_encontradas[0] + 2  # +2 para compensar cabeçalho e índice 0
        
        # Determinar a coluna para preenchimento
        coluna = None
        if pagador.lower() in ['alex-rute', 'alex rute', 'alex', 'rute']:
            coluna = 'E'  # Coluna Alex-Rute
        elif pagador.lower() in ['diego-ana', 'diego ana', 'diego', 'ana']:
            coluna = 'F'  # Coluna Diego-Ana
        else:
            # Tentar inferir com base no nome extraído do comprovante
            if 'alex' in pagador.lower() or 'rute' in pagador.lower():
                coluna = 'E'
            elif 'diego' in pagador.lower() or 'ana' in pagador.lower():
                coluna = 'F'
            else:
                print(f"Pagador '{pagador}' não reconhecido. Use 'Alex-Rute' ou 'Diego-Ana'.")
                coluna = input("Digite a coluna correta (E para Alex-Rute, F para Diego-Ana): ").upper()
                if coluna not in ['E', 'F']:
                    print("Coluna inválida.")
                    return False
        
        # Preencher o valor na célula apropriada
        self.sheet[f'{coluna}{linha_excel}'] = valor
        
        # Aplicar formatação verde (pago) à linha
        for col in range(1, 7):  # Colunas A a F
            coluna_letra = openpyxl.utils.get_column_letter(col)
            self.sheet[f'{coluna_letra}{linha_excel}'].fill = self.verde
        
        # Salvar as alterações
        self.workbook.save(self.excel_path)
        print(f"Pagamento de R$ {valor:.2f} registrado para '{atividade}' por {pagador}.")
        return True
    
    def atualizar_status(self):
        """Atualiza o status de pagamento de todas as linhas com base nos valores preenchidos"""
        # Para cada linha na tabela (exceto cabeçalho)
        for i in range(2, self.sheet.max_row + 1):
            valor_custo = self.sheet[f'A{i}'].value
            if valor_custo is None or not isinstance(valor_custo, (int, float)):
                continue
                
            alex_rute = self.sheet[f'E{i}'].value or 0
            diego_ana = self.sheet[f'F{i}'].value or 0
            
            # Verificar se o valor total foi pago
            valor_pago = 0
            if isinstance(alex_rute, (int, float)):
                valor_pago += alex_rute
            if isinstance(diego_ana, (int, float)):
                valor_pago += diego_ana
            
            # Definir cor com base no status de pagamento
            cor = self.verde if valor_pago >= valor_custo else self.vermelho
            
            # Aplicar cor à linha inteira
            for col in range(1, 7):  # Colunas A a F
                coluna_letra = openpyxl.utils.get_column_letter(col)
                self.sheet[f'{coluna_letra}{i}'].fill = cor
        
        # Salvar as alterações
        self.workbook.save(self.excel_path)
        print("Status de pagamento atualizado com sucesso.")
    
    def listar_atividades_pendentes(self):
            """Lista todas as atividades que ainda estão pendentes de pagamento, incluindo colunas E e F nulas"""
            atividades_pendentes = []
            
            # Para cada linha na tabela (exceto cabeçalho)
            for i in range(2, self.sheet.max_row + 1):
                valor_custo = self.sheet[f'B{i}'].value
                if valor_custo is None or not isinstance(valor_custo, (int, float)):
                    continue
                    
                alex_rute = self.sheet[f'E{i}'].value
                diego_ana = self.sheet[f'F{i}'].value
                
                # Verificar se as colunas E e F estão nulas
                if alex_rute is None and diego_ana is None:
                    atividade = self.sheet[f'C{i}'].value
                    setor = self.sheet[f'B{i}'].value
                    
                    atividades_pendentes.append({
                        'Atividade': atividade,
                        'Setor': setor,
                        'Valor Total': valor_custo
                    })
            
            return atividades_pendentes
    
    def listar_atividades(self):
        """Lista todas as atividades disponíveis na tabela"""
        atividades = []
        
        # Para cada linha na tabela (exceto cabeçalho)
        for i in range(2, self.sheet.max_row + 1):
            atividade = self.sheet[f'D{i}'].value
            setor = self.sheet[f'C{i}'].value
            valor_custo = self.sheet[f'B{i}'].value
            
            if atividade and valor_custo:
                # Garantir que o valor seja um número
                if isinstance(valor_custo, str):
                    try:
                        valor_custo = float(valor_custo.replace('R$', '').replace('.', '').replace(',', '.').strip())
                    except ValueError:
                        # Se não conseguir converter, manter como string
                        pass
                
                atividades.append({
                    'id': i-1,
                    'Atividade': atividade,
                    'Setor': setor,
                    'Valor': valor_custo
                })
        
        return atividades


class ComprovanteReader:
    @staticmethod
    def extrair_valor(texto):
        match = re.search(r'R?\$?\s?\d{1,3}(?:\.\d{3})*,\d{2}', texto)
        if match:
            return match.group()
        return None

    @staticmethod
    def extrair_data(texto):
        match = re.search(r'\d{2}/\d{2}/\d{4}', texto)
        if match:
            return match.group()
        return None

    @staticmethod
    def extrair_nome(texto):
        match = re.search(r'(?:Titular|Pagador|Quem pagou|Nome do titular)\s*[:\-]?\s*([A-Za-z\s]+)', texto, re.IGNORECASE)
        if match:
            nome = match.group(1).strip()
            nome = re.sub(r'\bCPF\b.*', '', nome, flags=re.IGNORECASE).strip()
            return ' '.join([word.capitalize() for word in nome.split()])

        match = re.search(r'\bde\s([A-Za-z\s]+)', texto, re.IGNORECASE)
        if match:
            nome = match.group(1).strip()
            nome = re.sub(r'\bCPF\b.*', '', nome, flags=re.IGNORECASE).strip()
            return ' '.join([word.capitalize() for word in nome.split()])

        return None

    @classmethod
    def ler_comprovante(cls, caminho_imagem):
        """Lê um comprovante de imagem e extrai as informações relevantes"""
        try:
            imagem = Image.open(caminho_imagem)
            texto = pytesseract.image_to_string(imagem, lang='eng')
            
            valor = cls.extrair_valor(texto)
            data = cls.extrair_data(texto)
            nome = cls.extrair_nome(texto)
            
            return {
                'texto_completo': texto,
                'valor': valor,
                'data': data,
                'nome': nome
            }
        except Exception as e:
            print(f"Erro ao processar a imagem: {e}")
            return None


def main():
    # Caminho para o arquivo Excel
    excel_file = "Fluxo Caixa Construção Guaratinguetá.xlsx"  # Altere para o caminho real do seu arquivo
    
    # Se o arquivo não existir, avisa o usuário
    if not os.path.exists(excel_file):
        print(f"Arquivo {excel_file} não encontrado. Verifique o caminho e tente novamente.")
        return
    
    # Inicializar o gerenciador de comprovantes
    manager = ComprovantesManager(excel_file)
    
    while True:
        print("\n==== SISTEMA DE GESTÃO DE GASTOS DE OBRA ====")
        print("1. Processar novo comprovante")
        print("2. Listar atividades pendentes")
        print("3. Atualizar status de pagamentos")
        print("4. Listar todas as atividades")
        print("0. Sair")
        
        opcao = input("\nEscolha uma opção: ")
        
        if opcao == "1":
            # Processar novo comprovante
            caminho_imagem = input("Digite o caminho da imagem do comprovante: ")
            if not os.path.exists(caminho_imagem):
                print(f"Arquivo {caminho_imagem} não encontrado.")
                continue
            
            # Ler o comprovante
            dados = ComprovanteReader.ler_comprovante(caminho_imagem)
            if not dados:
                print("Não foi possível processar o comprovante.")
                continue
            
            print("\nDados extraídos do comprovante:")
            print(f"Valor: {dados['valor'] or 'Não identificado'}")
            print(f"Data: {dados['data'] or 'Não identificada'}")
            print(f"Nome: {dados['nome'] or 'Não identificado'}")
            
            # Listar atividades disponíveis para seleção
            atividades = manager.listar_atividades()
            print("\nAtividades disponíveis:")
            for i, atv in enumerate(atividades):
                            valor = atv['Valor']
                            valor_str = ""
                            
                            # Formatar o valor corretamente
                            if isinstance(valor, (int, float)):
                                valor_str = f"R$ {valor:.2f}"
                            else:
                                valor_str = f"R$ {valor}"
                            
                            print(f"{i+1}. {atv['Atividade']} ({atv['Setor']}) - {valor_str}")
            
            # Obter inputs do usuário
            idx_atividade = int(input("\nSelecione o número da atividade: ")) - 1
            if idx_atividade < 0 or idx_atividade >= len(atividades):
                print("Índice de atividade inválido.")
                continue
            
            atividade_selecionada = atividades[idx_atividade]['Atividade']
            setor_selecionado = atividades[idx_atividade]['Setor']
            
            # Determinar pagador
            pagador = dados['nome'] if dados['nome'] else None
            if not pagador:
                pagador = input("Digite quem fez o pagamento (Alex-Rute ou Diego-Ana): ")
            else:
                confirmacao = input(f"Pagador identificado como '{pagador}'. Confirma? (S/N): ")
                if confirmacao.upper() != 'S':
                    pagador = input("Digite quem fez o pagamento (Alex-Rute ou Diego-Ana): ")
            
            # Preencher pagamento
            if dados['valor']:
                sucesso = manager.preencher_pagamento(
                    dados['valor'], 
                    atividade_selecionada, 
                    pagador, 
                    dados['data'], 
                    setor_selecionado
                )
                if sucesso:
                    # Atualizar status automaticamente após preencher
                    manager.atualizar_status()
            else:
                print("Valor não identificado no comprovante.")
                valor_manual = input("Digite o valor manualmente (formato: R$ X.XXX,XX): ")
                sucesso = manager.preencher_pagamento(
                    valor_manual, 
                    atividade_selecionada, 
                    pagador, 
                    dados['data'], 
                    setor_selecionado
                )
                if sucesso:
                    manager.atualizar_status()
            
        elif opcao == "2":
            # Listar atividades pendentes
            pendentes = manager.listar_atividades_pendentes()
            if not pendentes:
                print("Não há atividades pendentes de pagamento.")
            else:
                print("\n=== ATIVIDADES PENDENTES DE PAGAMENTO ===")
                for i, p in enumerate(pendentes):
                    print(f"{i+1}. {p['Atividade']} ({p['Setor']}): R$ {p['Valor Total']:.2f}")
                
        elif opcao == "3":
            # Atualizar status de pagamentos
            manager.atualizar_status()
        
        elif opcao == "4":
            # Listar todas as atividades
            atividades = manager.listar_atividades()
            print("\n=== TODAS AS ATIVIDADES ===")
            for i, atv in enumerate(atividades):
                        valor = atv['Valor']
                        valor_str = ""
                        
                        # Formatar o valor corretamente
                        if isinstance(valor, (int, float)):
                            valor_str = f"R$ {valor:.2f}"
                        else:
                            valor_str = f"R$ {valor}"
                        
                        print(f"{i+1}. {atv['Atividade']} ({atv['Setor']}) - {valor_str}")
        
        elif opcao == "0":
            print("Encerrando o programa...")
            break
        
        else:
            print("Opção inválida. Tente novamente.")


if __name__ == "__main__":
    # Verificar se o pytesseract está instalado corretamente
    try:
        import pytesseract
        # Definir caminho para o Tesseract (necessário no Windows)
        # Se estiver usando Windows, descomente a linha abaixo e ajuste o caminho
        # pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    except ImportError:
        print("Erro: pytesseract não está instalado.")
        print("Instale-o usando: pip install pytesseract")
        print("E instale o Tesseract OCR: https://github.com/tesseract-ocr/tesseract")
        sys.exit(1)
    
    main()