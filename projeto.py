from datetime import datetime
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
from PIL import Image
import pytesseract
import re
import io
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, Union
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Constants
UPLOAD_FOLDER = "uploads"
EXCEL_PATH = "planilha.xlsx"

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Pydantic models for request/response data
class Activity(BaseModel):
    id: int
    activity: str
    sector: Optional[str] = None
    value: float
    date: Optional[str] = None
    diego_ana: Optional[float] = None
    alex_rute: Optional[float] = None
    
class PendingActivity(BaseModel):
    id: int  
    activity: str
    sector: Optional[str] = None
    total_value: float
    valor_restante: float
    date: Optional[str] = None
    diego_ana: float
    alex_rute: float
    
class PaymentData(BaseModel):
    activity: str
    sector: Optional[str] = None
    payer: str
    value: str
    date: Optional[str] = None

class ExtractedData(BaseModel):
    value: Optional[str] = None
    date: Optional[str] = None
    name: Optional[str] = None
    full_text: Optional[str] = None


class ComprovanteReader:
    """Class responsible for reading and extracting information from payment receipts"""
    
    @staticmethod
    def extrair_valor(texto: str) -> Optional[str]:
        """Extract monetary value from text"""
        match = re.search(r'R?\$?\s?\d{1,3}(?:\.\d{3})*,\d{2}', texto)
        return match.group() if match else None

    @staticmethod
    def extrair_data(texto: str) -> Optional[str]:
        """Extract date in DD/MM/YYYY format from text"""
        match = re.search(r'\d{2}/\d{2}/\d{4}', texto)
        return match.group() if match else None

    @staticmethod
    def extrair_nome(texto: str) -> Optional[str]:
        """Extract person name from text"""
        # Try to find name after keywords like "Titular", "Pagador", etc.
        match = re.search(r'(?:Titular|Pagador|Quem pagou|Nome do titular)\s*[:\-]?\s*([A-Za-z\s]+)', texto, re.IGNORECASE)
        if match:
            nome = match.group(1).strip()
            nome = re.sub(r'\bCPF\b.*', '', nome, flags=re.IGNORECASE).strip()
            return ' '.join([word.capitalize() for word in nome.split()])

        # Alternative pattern: look for name after "de"
        match = re.search(r'\bde\s([A-Za-z\s]+)', texto, re.IGNORECASE)
        if match:
            nome = match.group(1).strip()
            nome = re.sub(r'\bCPF\b.*', '', nome, flags=re.IGNORECASE).strip()
            return ' '.join([word.capitalize() for word in nome.split()])

        return None

    @classmethod
    def ler_comprovante(cls, imagem_bytes: bytes) -> Dict[str, Any]:
        """Read a receipt from image bytes and extract relevant information"""
        try:
            with Image.open(io.BytesIO(imagem_bytes)) as imagem:
                texto = pytesseract.image_to_string(imagem, lang='eng')
                
                return {
                    'texto_completo': texto,
                    'valor': cls.extrair_valor(texto),
                    'data': cls.extrair_data(texto),
                    'nome': cls.extrair_nome(texto)
                }
        except Exception as e:
            logger.error(f"Error processing image: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error processing image: {str(e)}")


class ComprovantesManager:
    """Class to manage construction expenses in an Excel file"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        
        # Ensure the file exists
        if not os.path.exists(excel_path):
            self._create_excel_template()
            
        try:
            # Load the Excel file, maintaining existing formatting
            self.workbook = openpyxl.load_workbook(excel_path)
            self.sheet = self.workbook.active
            
            # Define colors for formatting
            self.verde = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            self.vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            
            # Load data to pandas for easier manipulation
            self.df = pd.read_excel(excel_path)
        except Exception as e:
            logger.error(f"Error initializing ComprovantesManager: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error loading Excel file: {str(e)}")
    
    def _create_excel_template(self) -> None:
        """Create a new Excel file with the expected structure"""
        wb = openpyxl.Workbook() 
        ws = wb.active
        
        # Set headers
        headers = ["Referência", "Valor", "Setor", "Atividade", "Alex-Rute", "Diego-Ana", "Data"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num).value = header
        
        # Save the file
        try:
            wb.save(self.excel_path)
        except Exception as e:
            logger.error(f"Error creating Excel template: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error creating Excel template: {str(e)}")
    
    def _find_activity_rows(self, activity: str, sector: Optional[str] = None) -> List[int]:
        """Find rows in Excel sheet that match the given activity and sector"""
        matching_rows = []
        
        for i in range(2, self.sheet.max_row + 1):
            cell_activity = self.sheet[f'D{i}'].value
            
            if cell_activity and cell_activity.strip().lower() == activity.strip().lower():
                matching_rows.append(i)
        
        # Filter by sector if provided
        if sector and matching_rows:
            sector_filtered = []
            for row in matching_rows:
                cell_sector = self.sheet[f'C{row}'].value
                if cell_sector and cell_sector.strip().lower() == sector.strip().lower():
                    sector_filtered.append(row)
            
            if sector_filtered:
                return sector_filtered
                
        return matching_rows

    def _parse_value(self, value_str: str) -> float:
        """Convert string value to float, handling different formats"""
        try:
            # Log para diagnóstico
            logger.debug(f"Parsing value: '{value_str}', type: {type(value_str)}")
            
            if isinstance(value_str, (int, float)):
                return float(value_str)
                    
            # Remover símbolos de moeda e espaços
            clean_value = value_str.replace('R$', '').strip()
            logger.debug(f"Clean value after removing currency symbol: '{clean_value}'")
            
            # Substituir separadores - formato brasileiro (1.234,56) para ponto decimal (1234.56)
            if ',' in clean_value:
                # Se tiver vírgula, assume formato brasileiro
                clean_value = clean_value.replace('.', '').replace(',', '.')
            
            logger.debug(f"Final clean value: '{clean_value}'")
            
            return float(clean_value)
        except ValueError as e:
            logger.error(f"Value parsing error for '{value_str}': {str(e)}")
            raise HTTPException(status_code=400, detail=f"Invalid value format: {value_str}")
    
    def _save_workbook(self) -> None:
        """Save Excel workbook with error handling"""
        try:
            self.workbook.save(self.excel_path)
            # Reload dataframe after saving
            self.df = pd.read_excel(self.excel_path)
        except PermissionError:
            logger.error(f"Permission error while saving Excel file: {self.excel_path}")
            raise HTTPException(
                status_code=500, 
                detail="Could not save Excel file. It may be open in another program."
            )
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error saving Excel file: {str(e)}")
    
    def _apply_row_format(self, row: int, color: PatternFill) -> None:
        """Apply formatting to a row"""
        for col in range(1, 7):  # Columns A to F
            col_letter = openpyxl.utils.get_column_letter(col)
            self.sheet[f'{col_letter}{row}'].fill = color
    
    def preencher_pagamento(self, valor_str: str, atividade: str, pagador: str, 
                            setor: Optional[str] = None, data: Optional[str] = None) -> Dict[str, Any]:
        """Register a payment in the Excel sheet"""
        # Convert value to float
        valor = self._parse_value(valor_str)
        
        # Find matching rows
        linhas_encontradas = self._find_activity_rows(atividade, setor)
        
        if not linhas_encontradas:
            raise HTTPException(status_code=404, detail=f"Activity '{atividade}' not found")
        
        # Use the first matching row
        linha_excel = linhas_encontradas[0]
        
        # Determine column for payment based on payer
        coluna = self._determine_payer_column(pagador)
        
        # Add value to existing value in cell
        valor_existente = self.sheet[f'{coluna}{linha_excel}'].value or 0
        if not isinstance(valor_existente, (int, float)):
            valor_existente = 0
            
        self.sheet[f'{coluna}{linha_excel}'] = valor_existente + valor
      
        
        # Apply green formatting (paid) to row
        self._apply_row_format(linha_excel, self.verde)
        
        # Save changes
        self._save_workbook()
        
        return {
            "sucesso": True,
            "mensagem": f"Pagamento no valor de R$ {valor:.2f} Registrado na atividade : '{atividade}' por {pagador}",
            "data": data
        }
    
    def _determine_payer_column(self, payer: str) -> str:
        """Determine the column letter based on payer name"""
        payer = payer.lower()
        
        if payer in ['alex-rute', 'alex rute', 'alex', 'rute']:
            return 'E'  # Alex-Rute column
        elif payer in ['diego-ana', 'diego ana', 'diego', 'ana']:
            return 'F'  # Diego-Ana column
        else:
            # Try to infer based on name extracted from receipt
            if 'alex' in payer or 'rute' in payer:
                return 'E'
            elif 'diego' in payer or 'ana' in payer:
                return 'F'
            else:
                raise HTTPException(
                    status_code=400, 
                    detail=f"Payer '{payer}' not recognized. Use 'Alex-Rute' or 'Diego-Ana'"
                )
        
    def atualizar_status(self) -> Dict[str, Any]:
        """Update payment status for all rows based on filled values"""
        updated_rows = 0
        
        for i in range(2, self.sheet.max_row + 1):
            valor_custo = self.sheet[f'B{i}'].value
            if valor_custo is None or not isinstance(valor_custo, (int, float)):
                continue
                
            alex_rute = self.sheet[f'E{i}'].value or 0
            diego_ana = self.sheet[f'F{i}'].value or 0
            
            # Check if total amount has been paid
            valor_pago = 0
            if isinstance(alex_rute, (int, float)):
                valor_pago += alex_rute
            if isinstance(diego_ana, (int, float)):
                valor_pago += diego_ana
            
            # Define color based on payment status
            cor = self.verde if valor_pago >= valor_custo else self.vermelho
            
            # Apply color to entire row
            self._apply_row_format(i, cor)
            
            updated_rows += 1
        
        # Save changes
        self._save_workbook()
        
        return {
            "success": True,
            "message": "Payment status updated successfully",
            "updated_rows": updated_rows
        }
    
    def listar_atividades_pendentes(self) -> List[PendingActivity]:
        """List all activities with pending payments"""
        atividades_pendentes = []

        for i in range(2, self.sheet.max_row + 1):
            valor_custo = self.sheet[f'B{i}'].value
            if valor_custo is None or not isinstance(valor_custo, (int, float)):
                continue
            
            alex_rute = self.sheet[f'E{i}'].value or 0
            diego_ana = self.sheet[f'F{i}'].value or 0

            # Calculate remaining amount to be paid
            valor_restante = valor_custo - (alex_rute + diego_ana)

            # Check if there's still a pending amount
            if valor_restante > 0:
                atividade = self.sheet[f'D{i}'].value
                setor = self.sheet[f'C{i}'].value
                data = self.sheet[f'A{i}'].value
                
                atividades_pendentes.append(PendingActivity(
                    id=i - 1,  # Use row index as ID
                    activity=atividade,
                    sector=setor,
                    total_value=valor_custo,
                    valor_restante=valor_restante,  
                    date=data.strftime("%d/%m/%Y") if hasattr(data, "strftime") else None,
                    alex_rute=alex_rute,
                    diego_ana=diego_ana
                ))

        return atividades_pendentes
    
    def listar_atividades(self) -> List[Activity]:
        """List all activities in the table"""
        atividades = []
        
        for i in range(2, self.sheet.max_row + 1):
            atividade = self.sheet[f'D{i}'].value
            setor = self.sheet[f'C{i}'].value
            valor_custo = self.sheet[f'B{i}'].value
            data = self.sheet[f'A{i}'].value
            diego_ana = self.sheet[f'F{i}'].value or 0
            alex_rute = self.sheet[f'E{i}'].value or 0
            
            if atividade and valor_custo:
                # Ensure value is a number
                if isinstance(valor_custo, str):
                    try:
                        valor_custo = float(valor_custo.replace('R$', '').replace('.', '').replace(',', '.').strip())
                    except ValueError:
                        # Skip this row if conversion fails
                        continue
                
                atividades.append(Activity(
                    id=i-1,
                    activity=atividade,
                    sector=setor,
                    value=valor_custo,
                    date=data.strftime("%d/%m/%Y") if hasattr(data, "strftime") else None,
                    alex_rute=alex_rute,
                    diego_ana=diego_ana
                ))
        
        return atividades
    
    def adicionar_atividade(self, data: Union[str, datetime], valor: float, 
                           setor: str, atividade: str) -> Dict[str, Any]:
        """Add a new activity to the table"""
        try:
            logger.debug(f"Adicionando atividade: data={data}, valor={valor}, setor={setor}, atividade={atividade}")
            
            # Determine next available row
            proxima_linha = self.sheet.max_row + 1
            
            # Convert date to correct format
            try:
                if isinstance(data, str):
                    # Check if format is YYYY-MM-DD (from HTML type="date" input)
                    if re.match(r'\d{4}-\d{2}-\d{2}', data):
                        data_formatada = datetime.strptime(data, "%Y-%m-%d")
                    else:
                        # Try DD/MM/YYYY format
                        data_formatada = datetime.strptime(data, "%d/%m/%Y")
                else:
                    data_formatada = data
            except ValueError as e:
                logger.error(f"Falha na conversão de data: {e}")
                # Use original value if conversion fails
                data_formatada = data
            
            # Ensure value is a float
            try:
                valor_float = float(valor)
            except (ValueError, TypeError):
                raise HTTPException(status_code=400, detail="O valor deve ser um número")
            
            # Add data to appropriate cells
            self.sheet[f'A{proxima_linha}'] = data_formatada
            self.sheet[f'B{proxima_linha}'] = valor_float  # Cost/Value
            self.sheet[f'C{proxima_linha}'] = setor  # Sector
            self.sheet[f'D{proxima_linha}'] = atividade  # Activity
            
            # Apply red formatting (pending) to row
            self._apply_row_format(proxima_linha, self.vermelho)
            
            # Save changes
            self._save_workbook()
            
            return {
                "success": True,
                "mensagem": f"Atividade: '{atividade}' adicionada com sucesso",
                "id": proxima_linha - 1,
                "atividade": atividade,
                "setor": setor,
                "valor": valor_float,
                "data": data
            }
        except HTTPException as he:
            # Re-raise HTTP exceptions
            raise he
        except Exception as e:
            logger.error(f"Erro inesperado ao adicionar a atividade: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Erro ao adcionar a atividade: {str(e)}")
            
    def calcular_valor_total(self) -> float:
        """Calculate total value of construction by summing activity values"""
        total_value = 0
        for i in range(2, self.sheet.max_row + 1):
            valor_custo = self.sheet[f'B{i}'].value
            if isinstance(valor_custo, (int, float)):
                total_value += valor_custo
        return total_value
    
    def calcular_valor_total_pago(self) -> float:
        """Calculate total amount paid by summing values in Alex-Rute and Diego-Ana columns"""
        total_pago = 0
        for i in range(2, self.sheet.max_row + 1):
            alex_rute = self.sheet[f'E{i}'].value or 0
            diego_ana = self.sheet[f'F{i}'].value or 0
            
            if isinstance(alex_rute, (int, float)):
                total_pago += alex_rute
            if isinstance(diego_ana, (int, float)):
                total_pago += diego_ana
                
        return total_pago
    
    def calcular_valor_pago_diego(self) -> float:
        """Calculate total amount paid by Diego-Ana"""
        total_diego = 0
        for i in range(2, self.sheet.max_row + 1):
            diego_ana = self.sheet[f'F{i}'].value or 0
            
            if isinstance(diego_ana, (int, float)):
                total_diego += diego_ana
                
        return total_diego
    
    def calcular_valor_pago_alex(self) -> float:
        """Calculate total amount paid by Alex-Rute"""
        total_alex = 0
        for i in range(2, self.sheet.max_row + 1):
            alex_rute = self.sheet[f'E{i}'].value or 0
            
            if isinstance(alex_rute, (int, float)):
                total_alex += alex_rute
                
        return total_alex


# Initialize FastAPI app
app = FastAPI(title="Construction Expense Manager API")

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Initialize the manager
manager = ComprovantesManager(EXCEL_PATH)

@app.get("/")
def read_root():
    return {"message": "Construction Expense Manager API"}

@app.get("/atividades", response_model=List[Activity])
def get_activities():
    try:
        return manager.listar_atividades()
    except Exception as e:
        logger.error(f"Error fetching activities: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error fetching activities: {str(e)}")

@app.get("/atividades-pendentes", response_model=List[PendingActivity])
def get_pending_activities():
    try:
        return manager.listar_atividades_pendentes()
    except Exception as e:
        logger.error(f"Error fetching pending activities: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error fetching pending activities: {str(e)}")

@app.post("/update-status")
def update_status():
    try:
        return manager.atualizar_status()
    except Exception as e:
        logger.error(f"Error updating status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error updating status: {str(e)}")

@app.post("/add-activity")
def add_activity(
    atividade: str = Form(...),
    valor: str = Form(...),
    setor: str = Form(...), 
    data: str = Form(...)
):
    logger.debug(f"Received data: activity={atividade}, value={valor}, sector={setor}, date={data}")
    
    try:
        # Convert string value to float, handling different number formats
        try:
            valor_float = float(valor.replace(',', '.'))
        except ValueError:
            raise HTTPException(status_code=400, detail="Value must be a number")
        
        result = manager.adicionar_atividade(data, valor_float, setor, atividade)
        logger.debug(f"Resultado da adição: {result}")
        return result
    except HTTPException as he:
        # Re-raise HTTP exceptions
        raise he
    except Exception as e:
        error_message = f"Erro no endpoint /add-activity: {str(e)}"
        logger.error(error_message, exc_info=True)
        raise HTTPException(status_code=500, detail=error_message)

@app.get("/valor-total")
def get_total_value():
    """Calculate total construction value by summing activity values"""
    try:
        total_value = manager.calcular_valor_total()
        return {"total": total_value}
    except Exception as e:
        logger.error(f"Erro ao calcular o valor total: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao calcular o valor total: {str(e)}")

@app.get("/valor-total-pago")
def get_valor_pago():
    """Calculate total amount paid by summing values in Alex-Rute and Diego-Ana columns"""
    try:
        total_pago = manager.calcular_valor_total_pago()
        return {"total_pago": total_pago}
    except Exception as e:
        logger.error(f"Erro ao calcular o valor total pago: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao calcular o valor total pago: {str(e)}")

@app.get("/valor-pago-diego")
def get_valor_pago_diego():
    try:
        total_pago_diego = manager.calcular_valor_pago_diego()
        return {"total_pago_diego": total_pago_diego}
    except Exception as e:
        logger.error(f"Erro ao calcular o total pago por diego-Ana : {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao calcular o total pago por diego-Ana :  {str(e)}")

@app.get("/valor-pago-alex")
def get_valor_pago_alex():
    try:
        total_pago_alex = manager.calcular_valor_pago_alex()
        return {"total_pago_alex": total_pago_alex}
    except Exception as e:
        logger.error(f"Erro ao calcular o total pago por Alex-Rute: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro ao calcular o total pago por Alex-Rute: {str(e)}")

@app.post("/process-receipt", response_model=ExtractedData)
async def process_receipt(file: UploadFile = File(...)):
    try:
        logger.debug(f"Arquivo recebido: {file.filename}, tipo: {file.content_type}")
        # Read uploaded file
        contents = await file.read()
        # Process receipt image
        result = ComprovanteReader.ler_comprovante(contents)
        # Return extracted data
        return ExtractedData(
            value=result['valor'],
            date=result['data'],
            name=result['nome'],
            full_text=result['texto_completo']
        )
    except Exception as e:
        logger.error(f"Erro ao processar o comprovante {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/register-payment")
def register_payment(payment: PaymentData):
    try:
        # Log detalhado para diagnóstico
        logger.debug(f"Received payment data: {payment.model_dump()}")
        
        # Validar os dados recebidos
        if not payment.activity:
            raise HTTPException(status_code=400, detail="Activity name is required")
        if not payment.payer:
            raise HTTPException(status_code=400, detail="Payer is required")
        if not payment.value:
            raise HTTPException(status_code=400, detail="Payment value is required")
            
        # Log do valor antes da conversão
        logger.debug(f"Payment value before processing: '{payment.value}'")
        
        # Register payment in spreadsheet (this might raise exceptions)
        try:
            result = manager.preencher_pagamento(
                payment.value,
                payment.activity,
                payment.payer,
                payment.sector,
                payment.date
            )
        except Exception as e:
            logger.error(f"Error in preencher_pagamento: {str(e)}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Error processing payment: {str(e)}")
        
        # Update status after registering payment
        try:
            status_result = manager.atualizar_status()
            logger.debug(f"Status update result: {status_result}")
        except Exception as e:
            logger.error(f"Error updating status: {str(e)}", exc_info=True)
            # Continue anyway since the payment was registered
        
        return {"message": "Payment registered successfully!", "result": result}
    except HTTPException as he:
        # Re-raise HTTP exceptions
        logger.error(f"HTTP Exception: {he.detail}")
        raise he
    except Exception as e:
        logger.error(f"Unhandled exception in register_payment: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error registering payment: {str(e)}")

# Run the app
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="localhost", port=8000)